import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime
import openpyxl
from openpyxl import Workbook
import os

# Excel File Setup
file_path = "Monthly_Bills.xlsx"

# Default Admin Password
default_admin_password = "admin002$"
new_admin_password = "janith005*"

# Check if the Excel file exists, if not create one
if not os.path.exists(file_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Home Up"  # Default to one account sheet
    sheet.append(["Account Number", "Month", "Bill Amount", "Paid Amount", "Remaining Balance", "Date", "Time"])
    workbook.save(file_path)


# Functions
def submit_data():
    account = account_var.get()
    bill_amount = bill_var.get()
    paid_amount = amount_paid_var.get()
    remaining_balance = remaining_balance_var.get()
    date = date_var.get()
    time = time_var.get()

    if not account or not bill_amount or not paid_amount:
        messagebox.showerror("Input Error", "All fields are required.")
        return

    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Check if the sheet for the account exists, create if not
    if account not in workbook.sheetnames:
        workbook.create_sheet(account)
        sheet = workbook[account]
        sheet.append(["Account Number", "Month", "Bill Amount", "Paid Amount", "Remaining Balance", "Date", "Time"])
    else:
        sheet = workbook[account]

    # Append data to the selected account's sheet
    sheet.append([account, datetime.now().strftime("%B %Y"), bill_amount, paid_amount, remaining_balance, date, time])
    workbook.save(file_path)
    messagebox.showinfo("Success", "Bill data saved successfully!")
    clear_fields()


def calculate_balance():
    try:
        bill = float(bill_var.get())
        paid = float(amount_paid_var.get())
        remaining = bill - paid
        remaining_balance_var.set(f"{remaining:.2f}")
    except ValueError:
        messagebox.showerror("Calculation Error", "Enter valid numeric values for bill and paid amounts.")


def clear_fields():
    account_var.set("")
    bill_var.set("")
    amount_paid_var.set("")
    remaining_balance_var.set("")
    date_var.set(datetime.now().strftime("%Y-%m-%d"))
    time_var.set(datetime.now().strftime("%H:%M:%S"))


def edit_excel():
    global default_admin_password  # Declare the variable as global before using it
    password = password_var.get()
    if password == default_admin_password:
        # Allow the user to change the password
        new_password = simpledialog.askstring("Change Password", "Enter new admin password:")
        if new_password:
            default_admin_password = new_password  # Modify the global password
            messagebox.showinfo("Password Changed", "Admin password has been changed.")
    else:
        messagebox.showerror("Access Denied", "Invalid admin password.")


def open_excel_file():
    password = password_var.get()
    if password == default_admin_password:
        os.system(f'start excel "{file_path}"')
    else:
        messagebox.showerror("Access Denied", "Invalid password.")


def show_result():
    password = password_var.get()
    if password == default_admin_password:
        os.system(f'start excel "{file_path}"')
    else:
        messagebox.showerror("Access Denied", "Invalid password.")


def upload_pdf():
    # File dialog to select the PDF or receipt
    file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    if file_path:
        messagebox.showinfo("File Uploaded", f"File {file_path} uploaded successfully!")


# GUI Setup
root = tk.Tk()
root.title("Monthly Bill Management System")
root.geometry("600x500")

# Variables
account_var = tk.StringVar()
bill_var = tk.StringVar()
amount_paid_var = tk.StringVar()
remaining_balance_var = tk.StringVar()
date_var = tk.StringVar()
time_var = tk.StringVar()
password_var = tk.StringVar()

# Initialize date and time
date_var.set(datetime.now().strftime("%Y-%m-%d"))
time_var.set(datetime.now().strftime("%H:%M:%S"))

# Layout
tk.Label(root, text="Account Number", font=("Arial", 12)).grid(row=0, column=0, padx=10, pady=5, sticky="w")
account_menu = ttk.Combobox(root, textvariable=account_var, values=[
    "4515215604 (Home Up)",
    "4510306709 (Home Down)",
    "4521245706 (Shop)",
    "4523059306 (Loku Land)",
    "4500310703 (Ahangama)"
], state="readonly")
account_menu.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Total Bill Amount", font=("Arial", 12)).grid(row=1, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=bill_var, font=("Arial", 12)).grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="Amount Paid", font=("Arial", 12)).grid(row=2, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=amount_paid_var, font=("Arial", 12)).grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="Remaining Balance", font=("Arial", 12)).grid(row=3, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=remaining_balance_var, font=("Arial", 12), state="readonly").grid(row=3, column=1, padx=10,
                                                                                              pady=5)

tk.Label(root, text="Date", font=("Arial", 12)).grid(row=4, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=date_var, font=("Arial", 12), state="readonly").grid(row=4, column=1, padx=10, pady=5)

tk.Label(root, text="Time", font=("Arial", 12)).grid(row=5, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=time_var, font=("Arial", 12), state="readonly").grid(row=5, column=1, padx=10, pady=5)

tk.Button(root, text="Calculate Balance", command=calculate_balance, font=("Arial", 12), bg="lightblue").grid(row=6,
                                                                                                              column=0,
                                                                                                              padx=10,
                                                                                                              pady=20)
tk.Button(root, text="Submit", command=submit_data, font=("Arial", 12), bg="lightgreen").grid(row=6, column=1, padx=10,
                                                                                              pady=20)

tk.Label(root, text="Admin Password", font=("Arial", 12)).grid(row=7, column=0, padx=10, pady=5, sticky="w")
tk.Entry(root, textvariable=password_var, font=("Arial", 12), show="*").grid(row=7, column=1, padx=10, pady=5)

tk.Button(root, text="Edit Excel", command=edit_excel, font=("Arial", 12), bg="orange").grid(row=8, column=0,
                                                                                             columnspan=2, pady=10)
tk.Button(root, text="Show Result", command=show_result, font=("Arial", 12), bg="yellow").grid(row=9, column=0,
                                                                                               columnspan=2, pady=10)
tk.Button(root, text="Upload PDF", command=upload_pdf, font=("Arial", 12), bg="lightpink").grid(row=10, column=0,
                                                                                                columnspan=2, pady=10)

# Run GUI
root.mainloop()
